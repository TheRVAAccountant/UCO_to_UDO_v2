import os
import sys
import subprocess
import logging
import pythoncom
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils import get_column_letter
import win32com.client as win32
from compare_ranges import main as compare_main
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from excel_utils import get_cell_value, get_calculated_value

def safe_convert_to_decimal(value, logger):
    """
    Safely convert a value to Decimal with error handling.
    """
    try:
        if value is None or value == "":
            return Decimal('0')  # Default to 0 if the value is None or empty

        # Check if the value is a string that starts with '=', indicating a formula
        if isinstance(value, str) and value.startswith('='):
            logger.error(f"Attempted to convert a formula string to Decimal: {value}")
            return Decimal('0')  # Or handle as needed

        # Convert result to Decimal if it's numeric or looks like a number
        return Decimal(str(value)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError) as e:
        logger.error(f"Invalid value for conversion to Decimal: {value} - Error: {e}")
        return Decimal('0')  # Return 0 for any invalid or unconvertible values

def open_excel_file(file_path, logger):
    """Open the Excel file using the default system application."""
    try:
        if os.name == 'nt':  # Windows
            os.startfile(file_path)
        elif os.name == 'posix':  # macOS and Linux
            opener = 'open' if sys.platform == 'darwin' else 'xdg-open'
            subprocess.call([opener, file_path])
        logger.info(f"Opened Excel file: {file_path}")
    except Exception as e:
        logger.error(f"Failed to open Excel file: {e}", exc_info=True)

def recalculate_workbook_in_excel(new_target_file, logger, progress_callback, retries=3):
    """Recalculate the workbook using Excel application."""
    from win32com.client import gencache, constants
    excel = None
    wb = None
    attempt = 0
    success = False
    try:
        # Initialize COM library with STA threading model
        pythoncom.CoInitializeEx(pythoncom.COINIT_APARTMENTTHREADED)
        
        # Create the Excel application object outside the loop
        excel = gencache.EnsureDispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False  # Suppress Excel alerts
        excel.AskToUpdateLinks = False  # Prevent prompts to update links
        excel.AlertBeforeOverwriting = False

        while attempt < retries and not success:
            try:
                attempt += 1
                logger.info(f"Attempt {attempt}: Opening Excel to recalculate and save the workbook: {new_target_file}")
                
                # Open the workbook with appropriate parameters
                wb = excel.Workbooks.Open(
                    new_target_file,
                    UpdateLinks=0,
                    ReadOnly=False,
                    IgnoreReadOnlyRecommended=True,
                    Notify=False
                )
                
                # Recalculate all open workbooks
                excel.CalculateFullRebuild()

                # Wait for calculations to complete
                while excel.CalculationState != constants.xlDone:
                    time.sleep(0.5)  # Wait half a second before checking again

                wb.Save()
                logger.info("Workbook recalculated and saved successfully in Excel.")
                progress_callback(25)
                success = True
            except Exception as e:
                logger.error(f"Attempt {attempt}: An error occurred while recalculating the workbook in Excel: {e}", exc_info=True)
                if attempt < retries:
                    logger.info(f"Retrying in 5 seconds... (Attempt {attempt + 1})")
                    time.sleep(5)
                else:
                    raise
            finally:
                # Ensure the workbook is closed properly
                if wb is not None:
                    try:
                        wb.Close(SaveChanges=False)
                    except Exception as e_close:
                        logger.error(f"Error closing workbook: {e_close}", exc_info=True)
                    wb = None  # Remove reference to the workbook
    except Exception as e:
        logger.error(f"Failed to recalculate the workbook after {retries} attempts: {e}", exc_info=True)
        raise
    finally:
        if excel is not None:
            excel.Quit()
            del excel  # Help garbage collection
        # Uninitialize COM library
        pythoncom.CoUninitialize()

def format_tickmark_cell(tickmark_cell, logger):
    """Apply formatting to the tickmark cell."""
    try:
        tickmark_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        tickmark_cell.font = Font(name="Calibri", color="FF0000", bold=True)
        tickmark_cell.alignment = Alignment(horizontal="center", vertical="center")
        logger.info(f"'Tickmark' added to cell {tickmark_cell.coordinate} with formatting applied.")
    except Exception as e:
        logger.error(f"An error occurred while formatting 'Tickmark' cell: {e}", exc_info=True)

from decimal import Decimal

def process_do_tb_sheet(target_wb, data_wb, certification_total, certification_sheet, total_cell, logger, progress_callback):
    """Process the DO TB sheet."""
    try:
        # Access the sheets from both workbooks
        sheet = target_wb["DO TB"]
        data_sheet = data_wb["DO TB"]
        logger.info("Processing 'DO TB' sheet.")

        values_to_find = ["422100", "422200"]
        found_values = {}
        first_occurrences = {}

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        number_format = '#,##0.00;[Red](#,##0.00)'
        border_style = Border(top=Side(border_style="thin"), bottom=Side(border_style="double"))

        logger.info("Searching for '422100' and '422200' in Column C of the 'DO TB' sheet.")
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=3, max_col=3):
            cell = row[0]
            cell_value = str(cell.value).strip()
            if cell_value in values_to_find and cell_value not in found_values:
                # Access calculated value from data_sheet
                data_cell = data_sheet.cell(row=cell.row, column=cell.column + 5)  # Column H (8)
                value_in_column_h = data_cell.value

                # Check if data_cell contains a formula string
                if isinstance(value_in_column_h, str) and value_in_column_h.startswith('='):
                    logger.error(f"Unexpected formula in data_sheet at row {cell.row}, column H: {value_in_column_h}")
                    value_in_column_h = None  # Optionally, skip or handle differently

                if value_in_column_h is not None:
                    # Convert to Decimal
                    decimal_value = safe_convert_to_decimal(value_in_column_h, logger)
                    found_values[cell_value] = decimal_value
                    first_occurrences[cell_value] = cell.row
                    logger.info(f"Found '{cell_value}' in row {cell.row}, value in Column H: {decimal_value}")

                    # Apply yellow fill to Column H in target workbook
                    col_h_cell = sheet.cell(row=cell.row, column=8)  # Column H is 8
                    col_h_cell.fill = yellow_fill

                    # Assign value to Column N (14) in target workbook
                    col_n_cell = sheet.cell(row=cell.row, column=14, value=decimal_value)
                    col_n_cell.number_format = number_format
                    logger.info(f"Value '{decimal_value}' added to Column N at row {cell.row}.")

                if len(found_values) == len(values_to_find):
                    break

        if len(found_values) < len(values_to_find):
            logger.error("One or both of '422100' or '422200' were not found in Column C.")
            return

        # Determine the rows where '422100' and '422200' were found
        row_422100 = first_occurrences["422100"]
        row_422200 = first_occurrences["422200"]
        sum_row = max(row_422100, row_422200) + 1

        # Assign sum formula to Column N in target workbook
        sum_formula = f"=N{row_422100}+N{row_422200}"
        sum_cell = sheet.cell(row=sum_row, column=14, value=sum_formula)
        sum_cell.font = Font(name="Calibri", size=11)
        sum_cell.border = border_style
        sum_cell.number_format = number_format
        logger.info(f"Sum formula '{sum_formula}' entered in Column N at row {sum_row}.")

        # Auto-fit Column N
        auto_fit_column(sheet, 'N', logger)

        # Calculate the sum directly in Python
        calculated_sum = sum(found_values.values())
        logger.info(f"Calculated sum directly in Python: {calculated_sum}")

        # Convert certification_total to Decimal if not already
        certification_total = safe_convert_to_decimal(certification_total, logger)

        # Compare the sums
        if abs(calculated_sum - certification_total) < Decimal('0.01'):
            add_tickmark(sheet, sum_row, 15, "8", "Wingdings 2", 10)
            add_tickmark(certification_sheet, total_cell.row + 1, 4, "a", "Marlett", 12)
            logger.info(f"Sums match. Tickmarks added.")
        else:
            add_x_mark(sheet, sum_row, 15)
            add_x_mark(certification_sheet, total_cell.row + 1, 4)
            logger.info(f"Sums do not match. X marks added.")

        progress_callback(75)
    except Exception as e:
        logger.error(f"An error occurred while processing the 'DO TB' sheet: {e}", exc_info=True)

def auto_fit_column(sheet, column_letter, logger):
    """Auto-fit the width of a column."""
    try:
        max_length = max(len(str(cell.value)) for cell in sheet[column_letter] if cell.value)
        adjusted_width = max_length + 2
        sheet.column_dimensions[column_letter].width = adjusted_width
        logger.info(f"Auto-fitted Column {column_letter} to width: {adjusted_width}")
    except Exception as e:
        logger.error(f"Error auto-fitting column {column_letter}: {e}")

def add_tickmark(sheet, row, col, value, font_name, font_size, is_match=True):
    """Add a tickmark to a cell with specified formatting and yellow fill for matching cells in the 'DO TB' sheet."""
    cell = sheet.cell(row=row, column=col, value=value)
    
    # Apply font based on whether the value is a match
    if is_match:
        cell.font = Font(name=font_name, size=font_size)
    else:
        cell.font = Font(name="Calibri", size=font_size, bold=True)
    
    # Apply alignment
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # If the sheet is 'DO TB' and is_match is True, fill the cell with yellow
    if sheet.title == "DO TB" and is_match:
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.fill = yellow_fill

def find_component_sheet(workbook, tab_name, tier_component_name, trading_partner_number):
    """Find the component sheet based on the given criteria."""
    for sheet_name in workbook.sheetnames:
        if tab_name in sheet_name:
            return workbook[sheet_name]
        elif tier_component_name in sheet_name:
            return workbook[sheet_name]
        elif str(trading_partner_number) in sheet_name:  # Convert to string here
            return workbook[sheet_name]
    return None

def add_x_mark(sheet, row, col):
    """Add an 'X' mark to a cell."""
    cell = sheet.cell(row=row, column=col, value="X")
    cell.font = Font(name="Calibri", size=11, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def process_certification_sheet(target_wb, data_wb, logger, progress_callback):
    """Process the Certification sheet and extract necessary information for comparison."""
    try:
        # Access the sheets from both workbooks
        sheet = target_wb["Certification"]
        data_sheet = data_wb["Certification"]
        logger.info("Processing 'Certification' sheet.")

        # Find 'Trading Partner Number' cell in the target workbook
        trading_partner_cell = None
        for row in sheet.iter_rows(max_col=1, max_row=sheet.max_row):
            if row[0].value == "Trading Partner Number":
                trading_partner_cell = row[0]
                break

        if not trading_partner_cell:
            logger.error("'Trading Partner Number' cell not found in 'Certification' sheet.")
            return None, None

        # Find 'Total ' cell in the target workbook
        total_cell = None
        for row in sheet.iter_rows(max_col=1, max_row=sheet.max_row):
            if row[0].value == "Total ":
                total_cell = row[0]
                break

        if not total_cell:
            logger.error("'Total ' cell not found in 'Certification' sheet.")
            return None, None

        # Access the calculated total from data_sheet
        data_total_cell = data_sheet.cell(row=total_cell.row, column=4)
        certification_total = safe_convert_to_decimal(data_total_cell.value, logger)
        logger.info(f"Certification total found: {certification_total} at row {total_cell.row}, Column D.")

        # Add 'Tickmark' to the target workbook's sheet
        tickmark_cell = sheet.cell(row=trading_partner_cell.row, column=8, value="Tickmark")
        format_tickmark_cell(tickmark_cell, logger)

        # Define the table range in the target workbook
        table_range = sheet[f"A{trading_partner_cell.row}:H{total_cell.row}"]
        data_table_range = data_sheet[f"A{trading_partner_cell.row}:H{total_cell.row}"]
        headers = [cell.value for cell in sheet[trading_partner_cell.row][0:8]]
        logger.info(f"Certification Table Headers: {headers}")

        # Extract detailed row data
        row_data = []
        for target_row, data_row in zip(table_range[1:], data_table_range[1:]):  # Skip header row
            trading_partner_number = data_row[0].value
            tier_component_name = data_row[1].value
            tab_name = data_row[6].value
            component_total_unfilled = safe_convert_to_decimal(data_row[3].value, logger)

            if all([trading_partner_number, tier_component_name, tab_name, component_total_unfilled]):
                row_data.append({
                    'trading_partner_number': str(trading_partner_number),
                    'tier_component_name': tier_component_name,
                    'tab_name': tab_name,
                    'component_total_unfilled': component_total_unfilled,
                    'row': target_row  # Use the row from the target workbook for modifications
                })

        # Call process_do_tb_sheet with the appropriate parameters
        process_do_tb_sheet(target_wb, data_wb, certification_total, sheet, total_cell, logger, progress_callback)

        progress_callback(50)
        return table_range, row_data

    except Exception as e:
        logger.error(f"An error occurred while processing the 'Certification' sheet: {e}", exc_info=True)
        return None, None

def process_uco_to_udo_sheet(target_wb, data_wb, component_name, logger, progress_callback):
    """Process the UCO to UDO sheet."""
    try:
        # Access the sheets from both workbooks
        sheet = target_wb["DO UCO to UDO"]
        data_sheet = data_wb["DO UCO to UDO"]
        logger.info("Processing 'DO UCO to UDO' sheet.")

        # Find 'Component' cell in the target workbook
        component_cell = None
        for row in sheet.iter_rows(max_col=1, max_row=sheet.max_row):
            if row[0].value == "Component":
                component_cell = row[0]
                break
        
        if not component_cell:
            logger.error("'Component' cell not found in 'DO UCO to UDO' sheet.")
            return None

        # Find '{component_name} Total' cell in the target workbook
        total_component_cell = None
        for row in sheet.iter_rows(max_col=1, max_row=sheet.max_row):
            if row[0].value == f"{component_name} Total":
                total_component_cell = row[0]
                break
        
        if not total_component_cell:
            logger.error(f"'{component_name} Total' cell not found in 'DO UCO to UDO' sheet.")
            return None

        # Add 'Tickmark' to the target workbook's sheet
        tickmark_cell = sheet.cell(row=component_cell.row, column=14, value="Tickmark")
        format_tickmark_cell(tickmark_cell, logger)

        # Evaluate and process the table range in the target workbook
        table_start_row = component_cell.row + 2
        table_end_row = total_component_cell.row
        table_range = sheet[f"A{table_start_row}:N{table_end_row}"]  # Columns A to N
        headers = [cell.value for cell in sheet[table_start_row][0:14]]  # Columns 1 to 14
        logger.info(f"DO UCO to UDO Table Headers: {headers}")

        # Process the relevant totals and convert them using safe_convert_to_decimal
        for target_row in sheet.iter_rows(min_row=table_start_row, max_row=table_end_row):
            row_num = target_row[0].row
            data_row_cells = data_sheet[row_num]

            # Access values from the data-only workbook
            uco_tier_component_name = data_row_cells[0].value  # Column A
            uco_component_total_unfilled = safe_convert_to_decimal(data_row_cells[4].value, logger)  # Column E (5th column)
            uco_trading_partner_total = safe_convert_to_decimal(data_row_cells[7].value, logger)     # Column H (8th column)
            uco_difference = safe_convert_to_decimal(data_row_cells[11].value, logger)               # Column L (12th column)

            logger.info(f"Row {row_num} - UCO Total: {uco_component_total_unfilled}, Trading Partner Total: {uco_trading_partner_total}, Difference: {uco_difference}")

            # Perform any necessary comparisons or processing
            # For example, you might compare these values with other data and add tickmarks accordingly

        progress_callback(60)
        return table_range

    except Exception as e:
        logger.error(f"An error occurred while processing 'DO UCO to UDO' sheet: {e}", exc_info=True)
        return None

def find_table_range(new_target_file, component_name, logger, progress_callback):
    """Main function to find table ranges, process sheets, and call comparison functions."""
    try:
        # Recalculate the workbook using Excel and refresh any external links or queries.
        recalculate_workbook_in_excel(new_target_file, logger, progress_callback)
        
        # Ensure a short delay to allow Excel to release the file
        time.sleep(1)

        # Load the workbook twice
        logger.info(f"Loading workbook with data_only=False: {new_target_file}")
        target_wb = load_workbook(new_target_file, data_only=False)  # Preserves formulas
        
        logger.info(f"Loading workbook with data_only=True: {new_target_file}")
        data_wb = load_workbook(new_target_file, data_only=True)     # Accesses calculated values

        # Process Certification sheet
        certification_range, certification_row_data = process_certification_sheet(target_wb, data_wb, logger, progress_callback)
        if certification_range is None or certification_row_data is None:
            logger.error("Failed to process Certification sheet. Aborting operation.")
            return
        
        # Process UCO to UDO sheet
        uco_to_udo_range = process_uco_to_udo_sheet(target_wb, data_wb, component_name, logger, progress_callback)
        if uco_to_udo_range is None:
            logger.error("Failed to process UCO to UDO sheet. Aborting operation.")
            return

        # Get the UCO to UDO sheet
        uco_to_udo_sheet = target_wb["DO UCO to UDO"]

        # Call the updated compare_main function to compare both UCO and UDO values
        if certification_range and uco_to_udo_range:
            compare_main(
                certification_range,
                uco_to_udo_range,
                target_wb,
                data_wb,
                logger,
                progress_callback,
                new_target_file
            )

        # Save the final workbook
        target_wb.save(new_target_file)
        logger.info(f"Workbook saved with updated tables and tickmark columns.")

        # Update progress after completion
        progress_callback(100)

        # Open the Excel file to show results to the user
        open_excel_file(new_target_file, logger)

    except InvalidFileException as e:
        logger.error(f"Invalid Excel file: {e}", exc_info=True)
    except Exception as e:
        logger.error(f"An unexpected error occurred: {e}", exc_info=True)

if __name__ == "__main__":
    # This block can be used for testing the script independently
    pass