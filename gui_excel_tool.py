import os
import logging
import shutil
import tkinter as tk
import threading
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageTk
from datetime import datetime
from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QLabel, QComboBox,
    QPushButton, QLineEdit, QFileDialog, QTextEdit, QHBoxLayout, QProgressBar
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QIcon
from find_table_range import find_table_range
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from background_worker import BackgroundWorker

def copy_cell_style(source_cell, target_cell):
    """Copies the style from source cell to target cell."""
    try:
        target_cell.font = source_cell.font.copy()
        target_cell.fill = source_cell.fill.copy()
        target_cell.border = source_cell.border.copy()
        target_cell.alignment = source_cell.alignment.copy()
        target_cell.number_format = source_cell.number_format
        target_cell.protection = source_cell.protection.copy()
    except Exception as e:
        raise RuntimeError(f"Failed to copy cell style: {e}")

def ensure_file_handle_release(file_path, logger):
    """
    Ensures Python releases the file handle before Excel operations
    
    Args:
        file_path (str): Path to the file that needs handle release
        logger (logging.Logger): Logger instance for operation tracking
    """
    try:
        import gc
        import time
        
        # Force Python garbage collection
        gc.collect()
        
        # Give Windows time to release file handles
        time.sleep(2)
        
        logger.info(f"Released file handle for: {file_path}")
    except Exception as e:
        logger.warning(f"Error during file handle release: {e}")

def copy_and_rename_sheet(source_path, source_sheet_name, target_path, new_sheet_name, logger, insert_index=None):
    try:
        logger.info(f"Loading source workbook: {source_path}")
        source_wb = load_workbook(source_path, data_only=False)  # data_only=False to preserve formulas
        if source_sheet_name not in source_wb.sheetnames:
            logger.error(f"Sheet '{source_sheet_name}' not found in {source_path}")
            return False

        logger.info(f"Loading target workbook: {target_path}")
        target_wb = load_workbook(target_path, data_only=False)  # Ensure formulas are preserved in the target too
        logger.info(f"Copying sheet '{source_sheet_name}' from source to target")
        source_sheet = source_wb[source_sheet_name]
        
        if insert_index is not None:
            target_sheet = target_wb.create_sheet(new_sheet_name, insert_index)
        else:
            target_sheet = target_wb.create_sheet(new_sheet_name)

        for row in source_sheet.iter_rows():
            for cell in row:
                target_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)

                # Preserve formulas
                if cell.data_type == "f":  # Check if the cell contains a formula
                    target_cell.value = cell.value

                # Copy style attributes
                if cell.has_style:
                    target_cell.font = cell.font.copy()
                    target_cell.border = cell.border.copy()
                    target_cell.fill = cell.fill.copy()
                    target_cell.number_format = cell.number_format
                    target_cell.protection = cell.protection.copy()
                    target_cell.alignment = cell.alignment.copy()

        # Copy column dimensions
        for key, value in source_sheet.column_dimensions.items():
            target_sheet.column_dimensions[key].width = value.width
            target_sheet.column_dimensions[key].hidden = value.hidden

        # Copy row dimensions
        for key, value in source_sheet.row_dimensions.items():
            target_sheet.row_dimensions[key].height = value.height
            target_sheet.row_dimensions[key].hidden = value.hidden

        logger.info(f"Saving changes to target workbook: {target_path}")
        target_wb.save(target_path)
        target_wb.close()
        source_wb.close()  # Make sure to close source workbook

        # Add file handle release after saving
        ensure_file_handle_release(target_path, logger)

        logger.info(f"Successfully copied and renamed sheet to '{new_sheet_name}' with formatting and formulas preserved")
        return True

    except Exception as e:
        logger.error(f"An error occurred while copying sheet: {e}", exc_info=True)
        return False

def create_copy_of_target_file(target_file, logger):
    """Creates a copy of the target file with a new name."""
    try:
        file_name, file_extension = os.path.splitext(target_file)
        new_file_name = f"{file_name} - DO{file_extension}"
        shutil.copy2(target_file, new_file_name)
        logger.info(f"Created copy of target file: {new_file_name}")

        # Add file handle release after copying
        ensure_file_handle_release(new_file_name, logger)
        
        return new_file_name
    except Exception as e:
        logger.error(f"Failed to create copy of target file: {e}")
        raise

class TextHandler(logging.Handler):
    def __init__(self, text_widget):
        super().__init__()
        self.text_widget = text_widget

    def emit(self, record):
        msg = self.format(record)
        self.text_widget.insert(tk.END, msg + '\n')
        self.text_widget.see(tk.END)
        self.text_widget.update_idletasks()

class MainWindow(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('UCO to UDO Recon')
        self.geometry('600x500')
        self.configure(bg='#313131')  # Set background color for the main window

        # Make the window resizable
        self.resizable(True, True)

        # Configure row and column weights
        self.grid_columnconfigure(1, weight=1)
        for i in range(8):  # Added one more row for cancel button
            self.grid_rowconfigure(i, weight=0)
        self.grid_rowconfigure(4, weight=1)  # Make the log text area expandable

        # Initialize the background worker
        self.background_worker = BackgroundWorker(
            update_progress_callback=self.update_progress,
            complete_callback=self.on_task_complete,
            error_callback=self.on_task_error
        )

        self.apply_forest_dark_theme()
        self.initUI()

        icon_path = "diamond_icon.ico"
        self.iconbitmap(icon_path)

        # Bind the window close event
        self.protocol("WM_DELETE_WINDOW", self.on_closing)

    def apply_forest_dark_theme(self):
        script_dir = os.path.dirname(os.path.abspath(__file__))
        tcl_file_path = os.path.join(script_dir, 'forest-dark.tcl')
        self.tk.call('source', tcl_file_path)
        ttk.Style().theme_use('forest-dark')

    def initUI(self):
        # Component Name
        ttk.Label(self, text="Select Component Name:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.component_name_combo = ttk.Combobox(self, values=["CBP", "CG", "CIS", "CYB", "FEM", "FLE", "ICE", "MGA", "MGT", "OIG", "TSA", "SS", "ST", "WMD"], state="readonly")
        self.component_name_combo.set("WMD")
        self.component_name_combo.grid(row=0, column=1, columnspan=2, sticky="ew", padx=5, pady=5)

        # UCO to UDO Reconciliation File
        ttk.Label(self, text="UCO to UDO Reconciliation File:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        self.target_file_edit = ttk.Entry(self)
        self.target_file_edit.grid(row=1, column=1, sticky="ew", padx=5, pady=5)
        ttk.Button(self, text="Browse...", command=lambda: self.browse_file(self.target_file_edit)).grid(row=1, column=2, padx=5, pady=5)

        # Trial Balance File
        ttk.Label(self, text="Trial Balance File:").grid(row=2, column=0, sticky="w", padx=5, pady=5)
        self.trial_balance_edit = ttk.Entry(self)
        self.trial_balance_edit.grid(row=2, column=1, sticky="ew", padx=5, pady=5)
        ttk.Button(self, text="Browse...", command=lambda: self.browse_file(self.trial_balance_edit)).grid(row=2, column=2, padx=5, pady=5)

        # UCO to UDO TIER File
        ttk.Label(self, text="UCO to UDO TIER File:").grid(row=3, column=0, sticky="w", padx=5, pady=5)
        self.uco_to_udo_edit = ttk.Entry(self)
        self.uco_to_udo_edit.grid(row=3, column=1, sticky="ew", padx=5, pady=5)
        ttk.Button(self, text="Browse...", command=lambda: self.browse_file(self.uco_to_udo_edit)).grid(row=3, column=2, padx=5, pady=5)

        # Modify the log text area to expand
        self.log_text = tk.Text(self, wrap=tk.WORD, width=70, height=10, bg='#232323', fg='white')
        self.log_text.grid(row=4, column=0, columnspan=3, padx=5, pady=5, sticky="nsew")
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.log_text.yview)
        scrollbar.grid(row=4, column=3, sticky="ns")
        self.log_text.configure(yscrollcommand=scrollbar.set)

        # Modify the progress bar to expand horizontally
        self.progress_bar = ttk.Progressbar(self, orient="horizontal", length=580, mode="determinate")
        self.progress_bar.grid(row=5, column=0, columnspan=3, padx=5, pady=5, sticky="ew")

        # Button frame with Start and Cancel buttons
        button_frame = ttk.Frame(self)
        button_frame.grid(row=6, column=0, columnspan=3, padx=5, pady=5, sticky="ew")
        button_frame.columnconfigure(0, weight=1)
        button_frame.columnconfigure(1, weight=1)

        # Start Button
        self.start_button = ttk.Button(button_frame, text="Start", command=self.start_operations)
        self.start_button.grid(row=0, column=0, padx=5, pady=5, sticky="e")

        # Cancel Button
        self.cancel_button = ttk.Button(button_frame, text="Cancel", command=self.cancel_operations, state="disabled")
        self.cancel_button.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        # Status label
        self.status_label = ttk.Label(self, text="Ready")
        self.status_label.grid(row=7, column=0, columnspan=3, padx=5, pady=5, sticky="w")

        # Configure grid
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(4, weight=1)

        # Set up logging
        self.logger = self.setup_logging()

    def browse_file(self, entry_widget):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if filename:
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, filename)

    def start_operations(self):
        """Start the main operations in a background thread with proper state management."""
        component_name = self.component_name_combo.get()
        target_file = self.target_file_edit.get()
        trial_balance_file = self.trial_balance_edit.get()
        uco_to_udo_file = self.uco_to_udo_edit.get()

        if not all([component_name, target_file, trial_balance_file, uco_to_udo_file]):
            messagebox.showerror("Error", "Please select all required files and component name.")
            return

        # Reset UI
        self.progress_bar['value'] = 0
        self.update_idletasks()
        self.status_label.config(text="Processing...")

        # Disable start button, enable cancel button
        self.start_button.config(state="disabled")
        self.cancel_button.config(state="normal")

        # Start background worker
        self.logger.info("Operation started...")
        try:
            self.background_worker.run_task(
                self.run_processing_task,
                component_name=component_name,
                target_file=target_file,
                trial_balance_file=trial_balance_file,
                uco_to_udo_file=uco_to_udo_file
            )
        except Exception as e:
            self.logger.error(f"Failed to start background worker: {e}", exc_info=True)
            self.on_task_error(e)

    def run_processing_task(self, component_name, target_file, trial_balance_file, uco_to_udo_file, cancel_event):
        """Main processing task to run in background."""
        try:
            # Check for cancellation
            if cancel_event.is_set():
                self.logger.info("Operation cancelled during startup.")
                return None

            # Create copy of target file
            new_target_file = create_copy_of_target_file(target_file, self.logger)
            ensure_file_handle_release(new_target_file, self.logger)

            # Check for cancellation after file copy
            if cancel_event.is_set():
                self.logger.info("Operation cancelled after file copy.")
                return None

            # Copy DO TB sheet
            if not copy_and_rename_sheet(trial_balance_file, f"{component_name} Total", new_target_file, "DO TB", self.logger, insert_index=3):
                self.logger.error(f"Failed to copy sheet '{component_name} Total'.")
                raise ValueError(f"Sheet '{component_name} Total' not found in Trial Balance file.")
            ensure_file_handle_release(new_target_file, self.logger)

            # Check for cancellation after first sheet copy
            if cancel_event.is_set():
                self.logger.info("Operation cancelled after DO TB sheet copy.")
                return None

            # Copy DO UCO to UDO sheet
            if not copy_and_rename_sheet(uco_to_udo_file, "UCO to UDO", new_target_file, "DO UCO to UDO", self.logger, insert_index=4):
                self.logger.error("Failed to copy 'UCO to UDO' sheet.")
                raise ValueError("Sheet 'UCO to UDO' not found in UCO to UDO TIER file.")
            ensure_file_handle_release(new_target_file, self.logger)

            # Update progress
            self.background_worker.update_progress(10)

            # Check for cancellation before main processing
            if cancel_event.is_set():
                self.logger.info("Operation cancelled before main processing.")
                return None

            # Perform the main operation with cancellation support
            find_table_range(
                new_target_file,
                component_name,
                self.logger,
                self.background_worker.update_progress,
                cancel_event
            )

            self.logger.info("Operations completed successfully.")
            return new_target_file  # Return the path to the created file

        except Exception as e:
            self.logger.error(f"Error during operation: {e}", exc_info=True)
            raise

    def cancel_operations(self):
        """Cancel the ongoing operation."""
        if self.background_worker.is_running():
            if messagebox.askyesno("Cancel Operation", "Are you sure you want to cancel the current operation?"):
                self.logger.info("User requested operation cancellation.")
                self.background_worker.cancel()
                self.status_label.config(text="Cancelling...")

    def on_task_complete(self, result):
        """Handle successful completion of the background task."""
        self.progress_bar['value'] = 100
        self.status_label.config(text="Complete")
        self.start_button.config(state="normal")
        self.cancel_button.config(state="disabled")

        if result:  # If we have a valid result (file path)
            self.logger.info(f"Operation completed successfully. Output file: {result}")
            messagebox.showinfo("Complete", f"Operations completed successfully!\nOutput file: {os.path.basename(result)}")

    def on_task_error(self, error):
        """Handle errors from the background task."""
        self.progress_bar['value'] = 0
        self.status_label.config(text="Error")
        self.start_button.config(state="normal")
        self.cancel_button.config(state="disabled")

        error_message = str(error)
        self.logger.error(f"Error during operation: {error_message}")
        messagebox.showerror("Error", f"An error occurred:\n{error_message}")

    def on_closing(self):
        """Handle window closing event with cleanup."""
        if self.background_worker.is_running():
            if messagebox.askyesno("Cancel Operation",
                               "An operation is still running. Are you sure you want to exit?"):
                self.logger.info("Closing application while operation is running.")
                self.background_worker.cancel()
                self.destroy()
        else:
            self.logger.info("Application closed normally.")
            self.destroy()

    def setup_logging(self):
        """Set up logging with file and text handlers."""
        logger = logging.getLogger("MainLogger")
        logger.setLevel(logging.DEBUG)

        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

        # File handler
        log_dir = "logs"
        os.makedirs(log_dir, exist_ok=True)
        log_filename = os.path.join(log_dir, f"UCOtoUDORecon_Log_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.txt")
        file_handler = logging.FileHandler(log_filename)
        file_handler.setLevel(logging.DEBUG)
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)

        # GUI Text handler
        text_handler = TextHandler(self.log_text)
        text_handler.setLevel(logging.DEBUG)
        text_handler.setFormatter(formatter)
        logger.addHandler(text_handler)

        return logger

    def update_progress(self, value):
        """Update the progress bar value with thread safety."""
        # Use after() to update from the main thread
        self.after(0, lambda: self._safe_update_progress(value))

    def _safe_update_progress(self, value):
        """Thread-safe progress bar update."""
        self.progress_bar['value'] = value
        self.update_idletasks()
    
def main():
    app = MainWindow()
    app.mainloop()

if __name__ == "__main__":
    main()
