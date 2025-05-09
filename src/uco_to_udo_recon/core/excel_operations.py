"""
Core Excel operations for the UCO to UDO Reconciliation tool.

This module provides functions for Excel file operations, including
copying worksheets, recalculating workbooks, and handling Excel automation.
"""

import os
import logging
import shutil
import time
import pythoncom
from typing import Optional, Union, Callable, Any
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell

from src.uco_to_udo_recon.utils.file_utils import ensure_file_handle_release


def copy_cell_style(source_cell: Cell, target_cell: Cell) -> None:
    """
    Copies the style from source cell to target cell.
    
    Args:
        source_cell: Source cell to copy style from
        target_cell: Target cell to apply style to
        
    Raises:
        RuntimeError: If copying cell style fails
    """
    try:
        target_cell.font = source_cell.font.copy()
        target_cell.fill = source_cell.fill.copy()
        target_cell.border = source_cell.border.copy()
        target_cell.alignment = source_cell.alignment.copy()
        target_cell.number_format = source_cell.number_format
        target_cell.protection = source_cell.protection.copy()
    except Exception as e:
        raise RuntimeError(f"Failed to copy cell style: {e}")


def create_copy_of_target_file(target_file: str, logger: logging.Logger) -> str:
    """
    Creates a copy of the target file with a new name.
    
    Args:
        target_file: Path to the target file to copy
        logger: Logger instance for tracking operations
        
    Returns:
        str: Path to the newly created file
        
    Raises:
        Exception: If file creation fails
    """
    try:
        file_name, file_extension = os.path.splitext(target_file)
        new_file_name = f"{file_name} - DO{file_extension}"
        shutil.copy2(target_file, new_file_name)
        logger.info(f"Created copy of target file: {new_file_name}")

        # Add file handle release after copying
        ensure_file_handle_release(new_file_name, logger)
        
        return new_file_name
    except Exception as e:
        logger.error(f"Failed to create copy of target file: {e}")
        raise


def copy_and_rename_sheet(
    source_path: str, 
    source_sheet_name: str, 
    target_path: str, 
    new_sheet_name: str, 
    logger: logging.Logger, 
    insert_index: Optional[int] = None,
    cancellation_check: Optional[Callable[[], bool]] = None
) -> bool:
    """
    Copies a sheet from source workbook to target workbook and renames it.
    
    Args:
        source_path: Path to the source Excel file
        source_sheet_name: Name of the sheet to copy
        target_path: Path to the target Excel file
        new_sheet_name: New name for the copied sheet
        logger: Logger instance for tracking operations
        insert_index: Optional index position to insert the sheet
        cancellation_check: Optional function to check if operation should be cancelled
        
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        # Check for cancellation
        if cancellation_check and cancellation_check():
            logger.info(f"Sheet copying cancelled for '{source_sheet_name}'.")
            return False
            
        logger.info(f"Loading source workbook: {source_path}")
        source_wb = load_workbook(source_path, data_only=False)  # data_only=False to preserve formulas
        if source_sheet_name not in source_wb.sheetnames:
            logger.error(f"Sheet '{source_sheet_name}' not found in {source_path}")
            return False

        logger.info(f"Loading target workbook: {target_path}")
        target_wb = load_workbook(target_path, data_only=False)  # Ensure formulas are preserved in the target too
        logger.info(f"Copying sheet '{source_sheet_name}' from source to target")
        source_sheet = source_wb[source_sheet_name]
        
        if insert_index is not None:
            target_sheet = target_wb.create_sheet(new_sheet_name, insert_index)
        else:
            target_sheet = target_wb.create_sheet(new_sheet_name)

        row_count = 0
        for row in source_sheet.iter_rows():
            # Check for cancellation periodically in large sheets
            row_count += 1
            if cancellation_check and cancellation_check() and row_count % 50 == 0:
                logger.info(f"Sheet copying cancelled during row processing for '{source_sheet_name}'.")
                return False
                
            for cell in row:
                target_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)

                # Preserve formulas
                if cell.data_type == "f":  # Check if the cell contains a formula
                    target_cell.value = cell.value

                # Copy style attributes
                if cell.has_style:
                    target_cell.font = cell.font.copy()
                    target_cell.border = cell.border.copy()
                    target_cell.fill = cell.fill.copy()
                    target_cell.number_format = cell.number_format
                    target_cell.protection = cell.protection.copy()
                    target_cell.alignment = cell.alignment.copy()

        # Check for cancellation before copying dimensions
        if cancellation_check and cancellation_check():
            logger.info(f"Sheet copying cancelled before copying dimensions for '{source_sheet_name}'.")
            return False
            
        # Copy column dimensions
        for key, value in source_sheet.column_dimensions.items():
            target_sheet.column_dimensions[key].width = value.width
            target_sheet.column_dimensions[key].hidden = value.hidden

        # Copy row dimensions
        for key, value in source_sheet.row_dimensions.items():
            target_sheet.row_dimensions[key].height = value.height
            target_sheet.row_dimensions[key].hidden = value.hidden

        logger.info(f"Saving changes to target workbook: {target_path}")
        target_wb.save(target_path)
        target_wb.close()
        source_wb.close()  # Make sure to close source workbook

        # Add file handle release after saving
        ensure_file_handle_release(target_path, logger)

        logger.info(f"Successfully copied and renamed sheet to '{new_sheet_name}' with formatting and formulas preserved")
        return True

    except Exception as e:
        logger.error(f"An error occurred while copying sheet: {e}", exc_info=True)
        return False


def recalculate_workbook_in_excel(
    file_path: str, 
    logger: logging.Logger, 
    progress_callback: Callable[[int, Optional[str]], None], 
    retries: int = 3,
    cancellation_check: Optional[Callable[[], bool]] = None
) -> None:
    """
    Recalculate the workbook using Excel application via COM automation.
    
    Args:
        file_path: Path to the Excel file
        logger: Logger instance for tracking operations
        progress_callback: Callback function to update progress (value, message)
        retries: Number of retry attempts (default: 3)
        cancellation_check: Optional function to check if operation should be cancelled
        
    Raises:
        Exception: If recalculation fails after retries
    """
    from win32com.client import gencache, constants
    excel = None
    wb = None
    attempt = 0
    success = False
    try:
        # Check for cancellation
        if cancellation_check and cancellation_check():
            logger.info("Workbook recalculation cancelled before starting.")
            return
            
        # Initialize COM library with STA threading model
        pythoncom.CoInitializeEx(pythoncom.COINIT_APARTMENTTHREADED)
        
        # Create the Excel application object outside the loop
        excel = gencache.EnsureDispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False  # Suppress Excel alerts
        excel.AskToUpdateLinks = False  # Prevent prompts to update links
        excel.AlertBeforeOverwriting = False

        while attempt < retries and not success:
            # Check for cancellation before each attempt
            if cancellation_check and cancellation_check():
                logger.info(f"Workbook recalculation cancelled before attempt {attempt+1}.")
                return
                
            try:
                attempt += 1
                logger.info(f"Attempt {attempt}: Opening Excel to recalculate and save the workbook: {file_path}")
                progress_callback(5 * attempt, f"Recalculating workbook (attempt {attempt})")
                
                # Open the workbook with appropriate parameters
                wb = excel.Workbooks.Open(
                    file_path,
                    UpdateLinks=0,
                    ReadOnly=False,
                    IgnoreReadOnlyRecommended=True,
                    Notify=False
                )
                
                # Recalculate all open workbooks
                excel.CalculateFullRebuild()

                # Wait for calculations to complete, with periodic cancellation checks
                check_counter = 0
                while excel.CalculationState != constants.xlDone:
                    time.sleep(0.5)  # Wait half a second before checking again
                    check_counter += 1
                    
                    # Check for cancellation every few loops
                    if cancellation_check and cancellation_check() and check_counter % 4 == 0:
                        logger.info("Workbook recalculation cancelled during calculation.")
                        return

                # Check for cancellation before saving
                if cancellation_check and cancellation_check():
                    logger.info("Workbook recalculation cancelled before saving.")
                    return
                    
                wb.Save()
                logger.info("Workbook recalculated and saved successfully in Excel.")
                progress_callback(25, "Workbook recalculated successfully")
                success = True
            except Exception as e:
                logger.error(f"Attempt {attempt}: An error occurred while recalculating the workbook in Excel: {e}", exc_info=True)
                if attempt < retries:
                    logger.info(f"Retrying in 5 seconds... (Attempt {attempt + 1})")
                    time.sleep(5)
                else:
                    raise
            finally:
                # Ensure the workbook is closed properly
                if wb is not None:
                    try:
                        wb.Close(SaveChanges=False)
                    except Exception as e_close:
                        logger.error(f"Error closing workbook: {e_close}", exc_info=True)
                    wb = None  # Remove reference to the workbook
    except Exception as e:
        logger.error(f"Failed to recalculate the workbook after {retries} attempts: {e}", exc_info=True)
        raise
    finally:
        if excel is not None:
            excel.Quit()
            del excel  # Help garbage collection
        # Uninitialize COM library
        pythoncom.CoUninitialize()