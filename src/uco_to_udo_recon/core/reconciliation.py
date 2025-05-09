"""
Core reconciliation logic for the UCO to UDO application.

This module contains the main functions for finding and comparing
data tables in Excel sheets for reconciliation.
"""

import logging
import time
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple, Union
from decimal import Decimal
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell

from src.uco_to_udo_recon.core.excel_operations import recalculate_workbook_in_excel
from src.uco_to_udo_recon.utils.excel_utils import (
    get_cell_value, get_calculated_value, safe_convert_to_decimal
)
from src.uco_to_udo_recon.utils.file_utils import open_excel_file


def format_tickmark_cell(tickmark_cell: Cell, logger: logging.Logger) -> None:
    """
    Apply formatting to the tickmark cell.
    
    Args:
        tickmark_cell: The cell to format
        logger: Logger instance for tracking operations
    """
    try:
        tickmark_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        tickmark_cell.font = Font(name="Calibri", color="FF0000", bold=True)
        tickmark_cell.alignment = Alignment(horizontal="center", vertical="center")
        logger.info(f"'Tickmark' added to cell {tickmark_cell.coordinate} with formatting applied.")
    except Exception as e:
        logger.error(f"An error occurred while formatting 'Tickmark' cell: {e}", exc_info=True)


def add_tickmark(
    sheet: Worksheet, 
    row: int, 
    col: int, 
    value: str, 
    font_name: str, 
    font_size: int, 
    is_match: bool = True
) -> None:
    """
    Add a tickmark to a cell with specified formatting.
    
    Args:
        sheet: The worksheet to add the tickmark to
        row: The row to add the tickmark to
        col: The column to add the tickmark to
        value: The tickmark value to insert
        font_name: The font name to use
        font_size: The font size to use
        is_match: Whether the tickmark indicates a match (True) or mismatch (False)
    """
    cell = sheet.cell(row=row, column=col, value=value)
    if is_match:
        cell.font = Font(name=font_name, size=font_size)
    else:
        cell.font = Font(name="Calibri", size=font_size, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Apply yellow fill only for the DO TB sheet
    if sheet.title == "DO TB":
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def add_x_mark(sheet: Worksheet, row: int, col: int) -> None:
    """
    Add an 'X' mark to a cell.
    
    Args:
        sheet: The worksheet to add the X mark to
        row: The row to add the X mark to
        col: The column to add the X mark to
    """
    cell = sheet.cell(row=row, column=col, value="X")
    cell.font = Font(name="Calibri", size=11, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_fit_column(sheet: Worksheet, column_letter: str, logger: logging.Logger) -> None:
    """
    Auto-fit the width of a column.
    
    Args:
        sheet: The worksheet containing the column
        column_letter: The letter of the column to auto-fit
        logger: Logger instance for tracking operations
    """
    try:
        max_length = max(len(str(cell.value)) for cell in sheet[column_letter] if cell.value)
        adjusted_width = max_length + 2
        sheet.column_dimensions[column_letter].width = adjusted_width
        logger.info(f"Auto-fitted Column {column_letter} to width: {adjusted_width}")
    except Exception as e:
        logger.error(f"Error auto-fitting column {column_letter}: {e}")


def find_component_sheet(
    workbook: Workbook, 
    tab_name: Optional[str], 
    tier_component_name: Optional[str], 
    trading_partner_number: Optional[Union[str, int]],
    logger: logging.Logger,
    cancellation_check: Optional[Callable[[], bool]] = None
) -> Optional[Worksheet]:
    """
    Find the component sheet using multiple fallback strategies with comprehensive component mappings.
    
    Args:
        workbook: The Excel workbook object
        tab_name: The tab name to search for (can be None)
        tier_component_name: The TIER component name (e.g., 'FEM', 'CBP')
        trading_partner_number: The trading partner number
        logger: Logger instance for tracking operations
        cancellation_check: Optional function to check if operation should be cancelled
    
    Returns:
        Worksheet object or None if not found
    """
    try:
        # Check for cancellation
        if cancellation_check and cancellation_check():
            logger.info("Component sheet search cancelled.")
            return None
            
        # Complete component mappings based on the Excel sheet
        component_mappings = {
            "CBP": ["CBP", "CBP-7005"],
            "CG": ["USCG", "CG", "USCG-7006"],
            "CIS": ["CIS", "CIS-7001"],
            "CYB": ["CISA", "CYB", "CISA-7009"],
            "FEM": ["FEMA", "FEM", "FEMA-7007"],
            "ICE": ["ICE", "ICE-7019"],
            "MGA": ["MGA", "MGA-7021"],
            "MGT": ["MGT", "MGT-7003"],
            "OIG": ["OIG", "OIG-7002"],
            "SS": ["USSS", "SS", "USSS-7004"],
            "ST": ["ST", "STA-7008"],
            "TSA": ["TSA", "TSA-7011"],
            "WMD": ["CWMD", "WMD", "CWMD-7023"]
        }

        # Trading partner mappings
        trading_partner_mappings = {
            "7005": ["CBP-7005"],
            "7006": ["USCG-7006"],
            "7001": ["CIS-7001"],
            "7009": ["CISA-7009"],
            "7007": ["FEMA-7007"],
            "7019": ["ICE-7019"],
            "7021": ["MGA-7021"],
            "7003": ["MGT-7003"],
            "7002": ["OIG-7002"],
            "7004": ["USSS-7004"],
            "7008": ["STA-7008"],
            "7011": ["TSA-7011"],
            "7023": ["CWMD-7023"]
        }

        # Normalize input values and create search patterns
        search_patterns = []
        
        # Build search patterns, handling None values
        if tab_name is not None:
            search_patterns.append((str(tab_name), "tab_name"))
        
        if tier_component_name is not None:
            # Add all possible component variations
            if tier_component_name in component_mappings:
                for variant in component_mappings[tier_component_name]:
                    search_patterns.append((variant, f"component_mapping_{tier_component_name}"))
            else:
                search_patterns.append((str(tier_component_name), "tier_component_name"))
        
        if trading_partner_number is not None:
            # Add trading partner variations
            str_trading_partner = str(trading_partner_number)
            search_patterns.append((str_trading_partner, "trading_partner_number"))
            if str_trading_partner in trading_partner_mappings:
                for variant in trading_partner_mappings[str_trading_partner]:
                    search_patterns.append((variant, f"trading_partner_mapping_{str_trading_partner}"))

        # List of sheets to skip
        skip_sheets = {
            "Instructions", 
            "Certification", 
            "DO TB", 
            "DO UCO to UDO"
        }

        # Search through all patterns
        for sheet_name in workbook.sheetnames:
            # Check for cancellation periodically
            if cancellation_check and cancellation_check():
                logger.info("Component sheet search cancelled.")
                return None
                
            # Skip certain sheet names
            if sheet_name in skip_sheets:
                continue
                
            sheet_name_upper = sheet_name.upper()
            
            for pattern, pattern_type in search_patterns:
                if pattern and pattern.upper() in sheet_name_upper:
                    logger.info(f"Found sheet '{sheet_name}' using {pattern_type} pattern: {pattern}")
                    return workbook[sheet_name]

        # If no match found, log detailed information
        logger.warning(
            f"No matching sheet found for TIER Component: {tier_component_name}\n"
            f"Search details:\n"
            f"- Tab Name: {tab_name}\n"
            f"- Trading Partner: {trading_partner_number}\n"
            f"- Search Patterns Used: {[p[0] for p in search_patterns]}\n"
            f"- Available sheets: {', '.join(workbook.sheetnames)}"
        )
        
        # Additional debugging information
        if tier_component_name:
            logger.debug(f"Component mappings available for {tier_component_name}: "
                        f"{component_mappings.get(tier_component_name, 'None')}")
        if trading_partner_number:
            logger.debug(f"Trading partner mappings available for {trading_partner_number}: "
                        f"{trading_partner_mappings.get(str(trading_partner_number), 'None')}")

        return None

    except Exception as e:
        logger.error(
            f"Error in find_component_sheet:\n"
            f"- Error message: {str(e)}\n"
            f"- TIER Component: {tier_component_name}\n"
            f"- Tab Name: {tab_name}\n"
            f"- Trading Partner: {trading_partner_number}\n"
            f"- Available sheets: {', '.join(workbook.sheetnames)}",
            exc_info=True
        )
        return None


def process_certification_sheet(
    target_wb: Workbook, 
    data_wb: Workbook, 
    logger: logging.Logger, 
    progress_callback: Callable[[int, Optional[str]], None],
    cancellation_check: Optional[Callable[[], bool]] = None
) -> Tuple[Optional[List[Any]], Optional[List[Dict[str, Any]]]]:
    """
    Process the Certification sheet and extract necessary information for comparison.
    
    Args:
        target_wb: The target workbook with formulas preserved
        data_wb: The data workbook with calculated values
        logger: Logger instance for tracking operations
        progress_callback: Callback function to update progress (value, message)
        cancellation_check: Optional function to check if operation should be cancelled
        
    Returns:
        Tuple containing the table range and row data, or (None, None) if processing fails
    """
    try:
        # Check for cancellation
        if cancellation_check and cancellation_check():
            logger.info("Certification sheet processing cancelled.")
            return None, None
            
        # Access the sheets from both workbooks
        sheet = target_wb["Certification"]
        data_sheet = data_wb["Certification"]
        logger.info("Processing 'Certification' sheet.")
        progress_callback(10, "Processing Certification sheet")

        # Find 'Trading Partner Number' cell in the target workbook
        trading_partner_cell = None
        for row in sheet.iter_rows(max_col=1, max_row=sheet.max_row):
            # Check for cancellation periodically
            if cancellation_check and cancellation_check():
                logger.info("Certification sheet processing cancelled.")
                return None, None
                
            if row[0].value == "Trading Partner Number":
                trading_partner_cell = row[0]
                break

        if not trading_partner_cell:
            logger.error("'Trading Partner Number' cell not found in 'Certification' sheet.")
            return None, None

        # Find 'Total ' cell in the target workbook
        total_cell = None
        for row in sheet.iter_rows(max_col=1, max_row=sheet.max_row):
            # Check for cancellation periodically
            if cancellation_check and cancellation_check():
                logger.info("Certification sheet processing cancelled.")
                return None, None
                
            if row[0].value == "Total ":
                total_cell = row[0]
                break

        if not total_cell:
            logger.error("'Total ' cell not found in 'Certification' sheet.")
            return None, None

        progress_callback(20, "Extracting certification data")

        # Access the calculated total from data_sheet
        data_total_cell = data_sheet.cell(row=total_cell.row, column=4)
        certification_total = safe_convert_to_decimal(data_total_cell.value, logger)
        logger.info(f"Certification total found: {certification_total} at row {total_cell.row}, Column D.")

        # Add 'Tickmark' to the target workbook's sheet
        tickmark_cell = sheet.cell(row=trading_partner_cell.row, column=8, value="Tickmark")
        format_tickmark_cell(tickmark_cell, logger)

        # Define the table range in the target workbook
        table_range = sheet[f"A{trading_partner_cell.row}:H{total_cell.row}"]
        data_table_range = data_sheet[f"A{trading_partner_cell.row}:H{total_cell.row}"]
        headers = [cell.value for cell in sheet[trading_partner_cell.row][0:8]]
        logger.info(f"Certification Table Headers: {headers}")

        progress_callback(30, "Preparing row data")

        # Extract detailed row data
        row_data = []
        for target_row, data_row in zip(table_range[1:], data_table_range[1:]):  # Skip header row
            # Check for cancellation periodically
            if cancellation_check and cancellation_check():
                logger.info("Certification sheet processing cancelled.")
                return None, None
                
            trading_partner_number = data_row[0].value
            tier_component_name = data_row[1].value
            tab_name = data_row[6].value
            component_total_unfilled = safe_convert_to_decimal(data_row[3].value, logger)

            if all([trading_partner_number, tier_component_name, tab_name, component_total_unfilled]):
                row_data.append({
                    'trading_partner_number': str(trading_partner_number),
                    'tier_component_name': tier_component_name,
                    'tab_name': tab_name,
                    'component_total_unfilled': component_total_unfilled,
                    'row': target_row  # Use the row from the target workbook for modifications
                })

        progress_callback(40, "Processing DO TB sheet")

        # Call process_do_tb_sheet with the appropriate parameters
        process_do_tb_sheet(target_wb, data_wb, certification_total, sheet, total_cell, logger, progress_callback, cancellation_check)

        progress_callback(50, "Certification sheet processing complete")
        return table_range, row_data

    except Exception as e:
        logger.error(f"An error occurred while processing the 'Certification' sheet: {e}", exc_info=True)
        return None, None


def process_do_tb_sheet(
    target_wb: Workbook, 
    data_wb: Workbook, 
    certification_total: Decimal, 
    certification_sheet: Worksheet, 
    total_cell: Cell, 
    logger: logging.Logger, 
    progress_callback: Callable[[int, Optional[str]], None],
    cancellation_check: Optional[Callable[[], bool]] = None
) -> None:
    """
    Process the DO TB sheet.
    
    Args:
        target_wb: The target workbook with formulas preserved
        data_wb: The data workbook with calculated values
        certification_total: The total from the certification sheet
        certification_sheet: The certification worksheet
        total_cell: The total cell in the certification sheet
        logger: Logger instance for tracking operations
        progress_callback: Callback function to update progress (value, message)
        cancellation_check: Optional function to check if operation should be cancelled
    """
    try:
        # Check for cancellation
        if cancellation_check and cancellation_check():
            logger.info("DO TB sheet processing cancelled.")
            return
            
        # Access the sheets from both workbooks
        sheet = target_wb["DO TB"]
        data_sheet = data_wb["DO TB"]
        logger.info("Processing 'DO TB' sheet.")
        progress_callback(55, "Processing DO TB sheet")

        values_to_find = ["422100", "422200"]
        found_values = {}
        first_occurrences = {}

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        number_format = '#,##0.00;[Red](#,##0.00)'
        border_style = Border(top=Side(border_style="thin"), bottom=Side(border_style="double"))

        logger.info("Searching for '422100' and '422200' in Column C of the 'DO TB' sheet.")
        progress_callback(60, "Searching for account codes")
        
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=3, max_col=3):
            # Check for cancellation periodically
            if cancellation_check and cancellation_check():
                logger.info("DO TB sheet processing cancelled.")
                return
                
            cell = row[0]
            cell_value = str(cell.value).strip()
            if cell_value in values_to_find and cell_value not in found_values:
                # Access calculated value from data_sheet
                data_cell = data_sheet.cell(row=cell.row, column=cell.column + 5)  # Column H (8)
                value_in_column_h = data_cell.value

                # Check if data_cell contains a formula string
                if isinstance(value_in_column_h, str) and value_in_column_h.startswith('='):
                    logger.error(f"Unexpected formula in data_sheet at row {cell.row}, column H: {value_in_column_h}")
                    value_in_column_h = None  # Optionally, skip or handle differently

                if value_in_column_h is not None:
                    # Convert to Decimal
                    decimal_value = safe_convert_to_decimal(value_in_column_h, logger)
                    found_values[cell_value] = decimal_value
                    first_occurrences[cell_value] = cell.row
                    logger.info(f"Found '{cell_value}' in row {cell.row}, value in Column H: {decimal_value}")

                    # Apply yellow fill to Column H in target workbook
                    col_h_cell = sheet.cell(row=cell.row, column=8)  # Column H is 8
                    col_h_cell.fill = yellow_fill

                    # Assign value to Column N (14) in target workbook
                    col_n_cell = sheet.cell(row=cell.row, column=14, value=decimal_value)
                    col_n_cell.number_format = number_format
                    logger.info(f"Value '{decimal_value}' added to Column N at row {cell.row}.")

                if len(found_values) == len(values_to_find):
                    break

        if len(found_values) < len(values_to_find):
            logger.error("One or both of '422100' or '422200' were not found in Column C.")
            return

        progress_callback(65, "Calculating summations")

        # Determine the rows where '422100' and '422200' were found
        row_422100 = first_occurrences["422100"]
        row_422200 = first_occurrences["422200"]
        sum_row = max(row_422100, row_422200) + 1

        # Assign sum formula to Column N in target workbook
        sum_formula = f"=N{row_422100}+N{row_422200}"
        sum_cell = sheet.cell(row=sum_row, column=14, value=sum_formula)
        sum_cell.font = Font(name="Calibri", size=11)
        sum_cell.border = border_style
        sum_cell.number_format = number_format
        logger.info(f"Sum formula '{sum_formula}' entered in Column N at row {sum_row}.")

        # Auto-fit Column N
        auto_fit_column(sheet, 'N', logger)

        # Calculate the sum directly in Python
        calculated_sum = sum(found_values.values())
        logger.info(f"Calculated sum directly in Python: {calculated_sum}")

        # Convert certification_total to Decimal if not already
        certification_total = safe_convert_to_decimal(certification_total, logger)

        # Compare the sums
        if abs(calculated_sum - certification_total) < Decimal('0.01'):
            add_tickmark(sheet, sum_row, 15, "8", "Wingdings 2", 10)
            add_tickmark(certification_sheet, total_cell.row + 1, 4, "a", "Marlett", 12)
            logger.info(f"Sums match. Tickmarks added.")
        else:
            add_x_mark(sheet, sum_row, 15)
            add_x_mark(certification_sheet, total_cell.row + 1, 4)
            logger.info(f"Sums do not match. X marks added.")

        progress_callback(75, "DO TB sheet processing complete")
    except Exception as e:
        logger.error(f"An error occurred while processing the 'DO TB' sheet: {e}", exc_info=True)


def process_uco_to_udo_sheet(
    target_wb: Workbook, 
    data_wb: Workbook, 
    component_name: str, 
    logger: logging.Logger, 
    progress_callback: Callable[[int, Optional[str]], None],
    cancellation_check: Optional[Callable[[], bool]] = None
) -> Optional[List[Any]]:
    """
    Process the UCO to UDO sheet.
    
    Args:
        target_wb: The target workbook with formulas preserved
        data_wb: The data workbook with calculated values
        component_name: The selected component name
        logger: Logger instance for tracking operations
        progress_callback: Callback function to update progress (value, message)
        cancellation_check: Optional function to check if operation should be cancelled
        
    Returns:
        The table range if successful, None otherwise
    """
    try:
        # Check for cancellation
        if cancellation_check and cancellation_check():
            logger.info("UCO to UDO sheet processing cancelled.")
            return None
            
        # Access the sheets from both workbooks
        sheet = target_wb["DO UCO to UDO"]
        data_sheet = data_wb["DO UCO to UDO"]
        logger.info("Processing 'DO UCO to UDO' sheet.")
        progress_callback(80, "Processing UCO to UDO sheet")

        # Find 'Component' cell in the target workbook
        component_cell = None
        for row in sheet.iter_rows(max_col=1, max_row=sheet.max_row):
            # Check for cancellation periodically
            if cancellation_check and cancellation_check():
                logger.info("UCO to UDO sheet processing cancelled.")
                return None
                
            if row[0].value == "Component":
                component_cell = row[0]
                break
        
        if not component_cell:
            logger.error("'Component' cell not found in 'DO UCO to UDO' sheet.")
            return None

        # Find '{component_name} Total' cell in the target workbook
        total_component_cell = None
        for row in sheet.iter_rows(max_col=1, max_row=sheet.max_row):
            # Check for cancellation periodically
            if cancellation_check and cancellation_check():
                logger.info("UCO to UDO sheet processing cancelled.")
                return None
                
            if row[0].value == f"{component_name} Total":
                total_component_cell = row[0]
                break
        
        if not total_component_cell:
            logger.error(f"'{component_name} Total' cell not found in 'DO UCO to UDO' sheet.")
            return None

        progress_callback(85, "Adding tickmarks")

        # Add 'Tickmark' to the target workbook's sheet
        tickmark_cell = sheet.cell(row=component_cell.row, column=14, value="Tickmark")
        format_tickmark_cell(tickmark_cell, logger)

        # Evaluate and process the table range in the target workbook
        table_start_row = component_cell.row + 2
        table_end_row = total_component_cell.row
        table_range = sheet[f"A{table_start_row}:N{table_end_row}"]  # Columns A to N
        headers = [cell.value for cell in sheet[table_start_row][0:14]]  # Columns 1 to 14
        logger.info(f"DO UCO to UDO Table Headers: {headers}")

        progress_callback(90, "Processing component totals")

        # Process the relevant totals and convert them using safe_convert_to_decimal
        for target_row in sheet.iter_rows(min_row=table_start_row, max_row=table_end_row):
            # Check for cancellation periodically
            if cancellation_check and cancellation_check():
                logger.info("UCO to UDO sheet processing cancelled.")
                return None
                
            row_num = target_row[0].row
            data_row_cells = data_sheet[row_num]

            # Access values from the data-only workbook
            uco_tier_component_name = data_row_cells[0].value  # Column A
            uco_component_total_unfilled = safe_convert_to_decimal(data_row_cells[4].value, logger)  # Column E (5th column)
            uco_trading_partner_total = safe_convert_to_decimal(data_row_cells[7].value, logger)     # Column H (8th column)
            uco_difference = safe_convert_to_decimal(data_row_cells[11].value, logger)               # Column L (12th column)

            logger.info(f"Row {row_num} - UCO Total: {uco_component_total_unfilled}, Trading Partner Total: {uco_trading_partner_total}, Difference: {uco_difference}")

        progress_callback(95, "UCO to UDO sheet processing complete")
        return table_range

    except Exception as e:
        logger.error(f"An error occurred while processing 'DO UCO to UDO' sheet: {e}", exc_info=True)
        return None


def find_table_range(
    new_target_file: str, 
    component_name: str, 
    logger: logging.Logger, 
    progress_callback: Callable[[int, Optional[str]], None],
    cancellation_check: Optional[Callable[[], bool]] = None
) -> None:
    """
    Main function to find table ranges, process sheets, and call comparison functions.
    
    Args:
        new_target_file: Path to the target Excel file
        component_name: The selected component name
        logger: Logger instance for tracking operations
        progress_callback: Callback function to update progress (value, message)
        cancellation_check: Optional function to check if operation should be cancelled
    """
    try:
        # Check for cancellation
        if cancellation_check and cancellation_check():
            logger.info("Table range processing cancelled.")
            return
            
        # Recalculate the workbook using Excel and refresh any external links or queries.
        progress_callback(5, "Recalculating workbook in Excel")
        recalculate_workbook_in_excel(new_target_file, logger, 
                                     lambda val: progress_callback(val, "Recalculating workbook"),
                                     cancellation_check=cancellation_check)
        
        # Check for cancellation after recalculation
        if cancellation_check and cancellation_check():
            logger.info("Table range processing cancelled after recalculation.")
            return
            
        # Ensure a short delay to allow Excel to release the file
        time.sleep(1)

        # Load the workbook twice
        progress_callback(30, "Loading workbooks")
        logger.info(f"Loading workbook with data_only=False: {new_target_file}")
        target_wb = load_workbook(new_target_file, data_only=False)  # Preserves formulas
        
        logger.info(f"Loading workbook with data_only=True: {new_target_file}")
        data_wb = load_workbook(new_target_file, data_only=True)     # Accesses calculated values

        # Check for cancellation after loading
        if cancellation_check and cancellation_check():
            logger.info("Table range processing cancelled after loading workbooks.")
            return

        # Process Certification sheet
        certification_range, certification_row_data = process_certification_sheet(
            target_wb, data_wb, logger, progress_callback, cancellation_check
        )
        if certification_range is None or certification_row_data is None:
            logger.error("Failed to process Certification sheet. Aborting operation.")
            return
        
        # Check for cancellation after processing certification sheet
        if cancellation_check and cancellation_check():
            logger.info("Table range processing cancelled after processing certification sheet.")
            return
        
        # Process UCO to UDO sheet
        uco_to_udo_range = process_uco_to_udo_sheet(
            target_wb, data_wb, component_name, logger, progress_callback, cancellation_check
        )
        if uco_to_udo_range is None:
            logger.error("Failed to process UCO to UDO sheet. Aborting operation.")
            return
            
        # Check for cancellation after processing UCO to UDO sheet
        if cancellation_check and cancellation_check():
            logger.info("Table range processing cancelled after processing UCO to UDO sheet.")
            return

        # Import here to avoid circular imports
        from src.uco_to_udo_recon.core.comparison import main as compare_main

        # Call the updated compare_main function to compare both UCO and UDO values
        if certification_range and uco_to_udo_range:
            compare_main(
                certification_range,
                uco_to_udo_range,
                target_wb,
                data_wb,
                logger,
                progress_callback,
                new_target_file,
                cancellation_check
            )

        # Check for cancellation after comparison
        if cancellation_check and cancellation_check():
            logger.info("Table range processing cancelled after comparison.")
            return

        # Save the final workbook
        progress_callback(98, "Saving workbook")
        target_wb.save(new_target_file)
        logger.info(f"Workbook saved with updated tables and tickmark columns.")

        # Update progress after completion
        progress_callback(100, "Process completed successfully")

        # Open the Excel file to show results to the user
        open_excel_file(new_target_file, logger)

    except InvalidFileException as e:
        logger.error(f"Invalid Excel file: {e}", exc_info=True)
    except Exception as e:
        logger.error(f"An unexpected error occurred: {e}", exc_info=True)