import logging
from openpyxl.styles import Font, Alignment, PatternFill
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import math
from excel_utils import get_cell_value, get_calculated_value

def convert_to_number(value):
    """Convert a value to a number, handling various formats."""
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    elif isinstance(value, str):
        try:
            # Remove commas and handle parentheses for negative numbers
            cleaned_value = value.replace(",", "").replace("(", "-").replace(")", "").strip()
            return Decimal(cleaned_value)
        except ValueError:
            return value
    return value

def safe_convert_to_decimal(value, logger):
    """
    Safely convert a value to Decimal with error handling.
    """
    try:
        if value is None or value == "":
            return Decimal('0')  # Default to 0 if the value is None or empty

        # Check if the value is a string that starts with '=', indicating a formula
        if isinstance(value, str) and value.startswith('='):
            logger.error(f"Attempted to convert a formula string to Decimal: {value}")
            return Decimal('0')  # Or handle as needed

        # Convert result to Decimal if it's numeric or looks like a number
        return Decimal(str(value)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError) as e:
        logger.error(f"Invalid value for conversion to Decimal: {value} - Error: {e}")
        return Decimal('0')  # Return 0 for any invalid or unconvertible values

def print_sample_comparison_rows(certification_values, uco_to_udo_values, logger, num_rows=5):
    """Print a sample of comparison rows for debugging."""
    logger.info(f"Printing first {num_rows} rows of comparison values:")
    for i in range(min(num_rows, len(certification_values), len(uco_to_udo_values))):
        cert_values = certification_values[i]
        uco_values = uco_to_udo_values[i]
        logger.info(f"Row {i+1} - Certification: {cert_values}, DO UCO to UDO: {uco_values}")

def compare_ranges(certification_range, uco_to_udo_range, target_wb, data_wb, logger, progress_callback, new_target_file, cancel_event=None):
    """
    Compare certification and UCO to UDO ranges, and always check UCO and UDO values in component sheets.

    Args:
        certification_range: Range of cells from the Certification sheet
        uco_to_udo_range: Range of cells from the DO UCO to UDO sheet
        target_wb: Workbook loaded with formulas preserved
        data_wb: Workbook loaded with calculated values
        logger: Logger instance
        progress_callback: Function to update progress
        new_target_file: Path to save the workbook
        cancel_event: Optional threading.Event to check for cancellation
    """
    try:
        logger.info("Starting comparison of Certification and DO UCO to UDO ranges.")

        certification_values = []
        uco_to_udo_values = []

        # Check for cancellation
        if cancel_event and cancel_event.is_set():
            logger.info("Operation cancelled at the start of comparison.")
            return

        # Step 1: Process UCO to UDO range first to collect all uco_tier_component_names
        uco_tier_component_names_set = set()
        for uco_row in uco_to_udo_range:
            # uco_row is a tuple of Cell objects
            uco_tier_component_name = uco_row[0].value  # Column A
            if uco_tier_component_name:
                uco_tier_component_names_set.add(uco_tier_component_name)

            uco_component_total_unfilled = safe_convert_to_decimal(
                data_wb["DO UCO to UDO"].cell(row=uco_row[0].row, column=5).value, logger
            )  # Column E
            uco_trading_partner_total = safe_convert_to_decimal(
                data_wb["DO UCO to UDO"].cell(row=uco_row[0].row, column=8).value, logger
            )    # Column H
            uco_difference = safe_convert_to_decimal(
                data_wb["DO UCO to UDO"].cell(row=uco_row[0].row, column=12).value, logger
            )           # Column L

            uco_to_udo_values.append(
                (uco_tier_component_name, uco_component_total_unfilled, uco_trading_partner_total, uco_difference, uco_row)
            )

        logger.info(f"Collected {len(uco_tier_component_names_set)} unique UCO Tier Component Names from UCO to UDO range.")

        # Check for cancellation
        if cancel_event and cancel_event.is_set():
            logger.info("Operation cancelled after processing UCO to UDO range.")
            return

        # Step 2: Process Certification range
        for cert_row in certification_range[1:]:  # Skip header row
            # cert_row is a tuple of Cell objects
            tier_component_name = cert_row[1].value  # Column B

            # Accessing data_wb to get calculated values
            row_number = cert_row[0].row  # Get the row number from the first cell
            component_total_unfilled = safe_convert_to_decimal(
                data_wb["Certification"].cell(row=row_number, column=4).value, logger
            )  # Column D
            trading_partner_total = safe_convert_to_decimal(
                data_wb["Certification"].cell(row=row_number, column=5).value, logger
            )    # Column E
            difference = safe_convert_to_decimal(
                data_wb["Certification"].cell(row=row_number, column=6).value, logger
            )              # Column F

            if not tier_component_name:
                logger.debug(f"Row {row_number}: 'tier_component_name' is empty. Skipping.")
                continue  # Skip rows where there's no component name

            # Ensure values are not None before applying abs()
            component_total_unfilled = component_total_unfilled if component_total_unfilled is not None else Decimal('0')
            trading_partner_total = trading_partner_total if trading_partner_total is not None else Decimal('0')
            difference = difference if difference is not None else Decimal('0')

            # Define conditions
            all_numeric_zero = (difference == Decimal('0') and
                                component_total_unfilled == Decimal('0') and
                                trading_partner_total == Decimal('0'))
            tier_not_in_uco = (tier_component_name not in uco_tier_component_names_set)

            # Combined condition: Skip if all numeric values are zero AND tier_component_name not in UCO range
            if all_numeric_zero and tier_not_in_uco:
                logger.debug(f"Row {row_number}: All numeric values are zero and 'tier_component_name' ({tier_component_name}) not found in UCO Tier Component Names. Skipping.")
                continue  # Skip rows based on the combined condition

            # If all conditions pass, add to certification_values
            certification_values.append(
                (tier_component_name, component_total_unfilled, trading_partner_total, difference, cert_row)
            )
            logger.debug(f"Row {row_number}: Added to certification_values.")

        logger.info(f"Total certification_values after applying conditions: {len(certification_values)}")

        # Check for cancellation
        if cancel_event and cancel_event.is_set():
            logger.info("Operation cancelled after processing Certification range.")
            return

        print_sample_comparison_rows(certification_values, uco_to_udo_values, logger)

        total_comparisons = len(certification_values)
        logger.info(f"Total comparisons to make: {total_comparisons}")

        for idx, cert_values in enumerate(certification_values):
            # Check for cancellation periodically
            if cancel_event and cancel_event.is_set():
                logger.info(f"Operation cancelled during component comparison (at component {idx+1}/{total_comparisons}).")
                return

            # Update progress based on components processed
            if total_comparisons > 0:
                current_progress = 60 + int((idx / total_comparisons) * 40)  # Scale from 60% to 100%
                progress_callback(current_progress)

            tier_component_name, component_total_unfilled, trading_partner_total, difference, cert_row = cert_values

            # Always check the component sheet for UCO and UDO totals
            component_sheet = find_component_sheet(target_wb, cert_row[6].value, tier_component_name, cert_row[0].value, logger)
            if component_sheet:
                logger.info(f"Processing component sheet: {component_sheet.title}")

                # UCO comparison as done previously
                uco_cell = next((cell for row in component_sheet.iter_rows() for cell in row if isinstance(cell.value, str) and "UCO total reported in TIER" in cell.value), None)
                if uco_cell:
                    # Log the cell reference
                    logger.debug(f"UCO cell found at row {uco_cell.row}, column {2} in '{component_sheet.title}' sheet.")

                    # Access calculated UCO value from data_wb
                    cell_value = component_sheet.cell(row=uco_cell.row, column=2).value
                    logger.debug(f"Raw UCO cell value: {cell_value}")
                    data_uco_value = safe_convert_to_decimal(cell_value, logger)  # Assuming column B
                    logger.info(f"Processed UCO value from component sheet: {data_uco_value}")

                    # Compare UCO value from component sheet with UCO to UDO value
                    uco_to_udo_row_match = next((row for row in uco_to_udo_values if row[0] == tier_component_name), None)
                    if uco_to_udo_row_match:
                        uco_to_udo_value = uco_to_udo_row_match[1]  # Column E: component_total_unfilled
                        is_match = abs(data_uco_value - uco_to_udo_value) < Decimal('0.01')
                        add_tickmark(component_sheet, uco_cell.row + 1, 2, "i" if is_match else "X", "Wingdings", 11, is_match)
                        logger.info(f"UCO: {data_uco_value} compared with UCO to UDO: {uco_to_udo_value} - {'Match' if is_match else 'No Match'}")
                        logger.info(f"Tickmark added to component sheet {component_sheet.title} for TIER Component: {tier_component_name}")
                    else:
                        logger.warning(f"No matching UCO value found in UCO to UDO sheet for {tier_component_name}.")
                else:
                    logger.warning(f"UCO total cell not found in component sheet {component_sheet.title} for {tier_component_name}")

                # Check for cancellation between UCO and UDO processing
                if cancel_event and cancel_event.is_set():
                    logger.info(f"Operation cancelled during component {tier_component_name} processing (after UCO).")
                    return

                # UDO comparison
                udo_cell = next((cell for row in component_sheet.iter_rows() for cell in row if isinstance(cell.value, str) and "UDO total reported in TIER" in cell.value), None)
                if udo_cell:
                    # Log the cell reference
                    logger.debug(f"UDO cell found at row {udo_cell.row}, column {4} in '{component_sheet.title}' sheet.")

                    # Access calculated UDO value from data_wb
                    cell_value = component_sheet.cell(row=udo_cell.row, column=4).value
                    logger.debug(f"Raw UDO cell value: {cell_value}")
                    data_udo_value = safe_convert_to_decimal(cell_value, logger)  # Assuming column D
                    logger.info(f"Processed UDO value from component sheet: {data_udo_value}")

                    # Compare UDO value with UCO to UDO's trading partner total
                    uco_to_udo_row_match = next((row for row in uco_to_udo_values if row[0] == tier_component_name), None)
                    if uco_to_udo_row_match:
                        uco_to_udo_trading_partner_value = uco_to_udo_row_match[2]  # Column H: trading_partner_total
                        is_match = abs(data_udo_value - uco_to_udo_trading_partner_value) < Decimal('0.01')
                        add_tickmark(component_sheet, udo_cell.row + 1, 4, "i" if is_match else "X", "Wingdings", 11, is_match)

                        # Check for cancellation before recon table processing
                        if cancel_event and cancel_event.is_set():
                            logger.info(f"Operation cancelled during component {tier_component_name} processing (before recon table).")
                            return

                        # Process the recon table and pass new_target_file for saving
                        process_recon_table(component_sheet, data_wb, logger, new_target_file, udo_cell.row)  # Updated call
                        logger.info(f"UDO: {data_udo_value} compared with UCO to UDO Trading Partner Total: {uco_to_udo_trading_partner_value} - {'Match' if is_match else 'No Match'}")
                        logger.info(f"UDO Tickmark added to component sheet {component_sheet.title} for TIER Component: {tier_component_name}")
                    else:
                        logger.warning(f"No matching UDO value found in UCO to UDO sheet for {tier_component_name}.")
                else:
                    logger.warning(f"UDO total cell not found in component sheet {component_sheet.title} for {tier_component_name}")

            # Check for cancellation before sheet match processing
            if cancel_event and cancel_event.is_set():
                logger.info(f"Operation cancelled during component {tier_component_name} matching.")
                return

            # Handle the match between Certification and DO UCO to UDO sheets
            for uco_values in uco_to_udo_values:
                uco_tier_component_name, uco_component_total_unfilled, uco_trading_partner_total, uco_difference, uco_row = uco_values

                if (tier_component_name == uco_tier_component_name and
                    abs(component_total_unfilled - uco_component_total_unfilled) < Decimal('0.01') and
                    abs(trading_partner_total - uco_trading_partner_total) < Decimal('0.01') and
                    abs(difference - uco_difference) < Decimal('0.01')):

                    logger.info(f"Match found for TIER Component Name: {tier_component_name}")

                    # Add Tickmark to Certification sheet
                    certification_sheet = target_wb["Certification"]
                    add_tickmark(certification_sheet, cert_row[7].row, cert_row[7].column, "i", "Wingdings", 12, True)

                    # Add Tickmark to DO UCO to UDO sheet
                    uco_to_udo_sheet = target_wb["DO UCO to UDO"]
                    add_tickmark(uco_to_udo_sheet, uco_row[13].row, uco_row[13].column, "8", "Wingdings 2", 12, True)

                    logger.info(f"Tickmarks added to Certification and DO UCO to UDO sheets for TIER Component Name: {tier_component_name}")

        # Check for cancellation before saving
        if cancel_event and cancel_event.is_set():
            logger.info("Operation cancelled before saving workbook.")
            return

        # After processing all comparisons, save the workbook once
        target_wb.save(new_target_file)
        logger.info(f"Workbook saved with updated comparisons and tickmarks.")

        # Update progress to 100%
        progress_callback(100)

    except Exception as e:
        logger.error(f"An error occurred during the comparison: {e}", exc_info=True)
        raise

def process_recon_table(component_sheet, data_wb, logger, new_target_file, udo_row):
    """
    Process the recon table in the component sheet.
    Finds the header row, total row, and adds formulas in the tickmark row and 
    System of Record vs TIER tickmark row.
    The formula will be applied to columns B, D, E, F, G, H, and I (skipping C).
    """
    try:
        # Find "Contract / Agreement / Sales Order #" in Column A (Header Row)
        header_row = None
        total_row = None
        tickmark_row = None
        system_of_record_row = None
        udo_total_system_row = None
        udo_after_adjustments_row = None
        udo_tickmark_row = None
        difference_adjustments_row = None
        difference_adjustments_tickmark_row = None

        # Iterate over actual Cell objects in Column A to find header and total rows
        for row in component_sheet.iter_rows(min_col=1, max_col=1):
            cell_value = row[0].value  # Get the value of the cell in column A
            if cell_value == "Contract / Agreement / Sales Order #":
                header_row = row[0].row  # Get the row number of the cell
                logger.info(f"Found 'Contract / Agreement / Sales Order #' in Column A at row {header_row}")

            if cell_value == "Providing Bureau UCO Total via their system records:":
                total_row = row[0].row  # Get the row number of the cell
                logger.info(f"Found 'Providing Bureau UCO Total via their system records:' in Column A at row {total_row}")

            if cell_value == "Difference between: System of Record vs TIER":
                system_of_record_row = row[0].row
                logger.info(f"Found 'Difference between: System of Record vs TIER' in Column A at row {system_of_record_row}")

        # Add additional logic for UDO-related rows in Column C
        for row in component_sheet.iter_rows(min_col=3, max_col=3):
            cell_value = row[0].value  # Get the value of the cell in column C
            logger.info(f"Inspecting cell in Column C, row {row[0].row}: {cell_value}")
        
            if cell_value == "UDO total via system records":
                udo_total_system_row = row[0].row  # Get the row number of the cell
                logger.info(f"Found 'UDO total via system records' in Column C at row {udo_total_system_row}")
        
            if cell_value == "UDO after high level adjustments":
                udo_after_adjustments_row = row[0].row  # Get the row number of the cell
                udo_tickmark_row = udo_after_adjustments_row + 1  # The tickmark row is the row after this one
                logger.info(f"Found 'UDO after high level adjustments' in Column C at row {udo_after_adjustments_row}")
        
            if cell_value and isinstance(cell_value, str) and cell_value.strip() == "Difference between: System of Record (after adjustments) vs TIER":
                difference_adjustments_row = row[0].row  # Get the row number of the cell
                difference_adjustments_tickmark_row = difference_adjustments_row + 1  # The tickmark row is the row after this one
                logger.info(f"Found 'Difference between: System of Record (after adjustments) vs TIER' in Column C at row {difference_adjustments_row}")
        
        if not header_row or not total_row or not system_of_record_row or not udo_total_system_row or not udo_after_adjustments_row or not difference_adjustments_row:
            logger.warning("Could not find the required rows in the recon table.")
            return

        # Insert a new column after Column J
        component_sheet.insert_cols(11)  # 11 corresponds to Column J (after J)
        logger.info(f"Inserted a new column after Column J.")
        
        # Adjust the width of the new column (K)
        component_sheet.column_dimensions['K'].width = 45
        logger.info(f"Set the width of column K to 45.")

        # Set the header for the new column in the header_row
        do_comments_cell = component_sheet.cell(row=header_row, column=11, value="DO Comments")

        # Apply formatting to the new header cell
        do_comments_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow fill
        do_comments_cell.font = Font(color="FF0000", bold=True, size=11, name="Calibri")  # Bold font
        do_comments_cell.alignment = Alignment(horizontal="center", vertical="center")  # Center and middle alignment

        logger.info(f"Added 'DO Comments' header to the new column at row {header_row} and applied formatting.")

        # Format the rest of the column with no fill, bold red font
        for row_idx in range(header_row + 1, component_sheet.max_row + 1):
            comment_cell = component_sheet.cell(row=row_idx, column=11)
            comment_cell.font = Font(color="FF0000", bold=True, size=11, name="Calibri")  # Red Calibri bold font, size 11
            comment_cell.fill = PatternFill(fill_type=None)  # No fill

        logger.info(f"Formatted the new 'DO Comments' column with red bold text.")


        # The first row of data is right after the header row
        first_data_row = header_row + 1

        # The last row of data is right before the total row
        last_data_row = total_row - 1

        # The tickmark row is right after the total row
        tickmark_row = total_row + 1

        logger.info(f"Recon Table Data range: Rows {first_data_row} to {last_data_row}")
        logger.info(f"Tickmark row is {tickmark_row}")

        # List of columns where the formula will be added (B to H => Columns 2, 4-8, skipping C which is index 3)
        columns = ["B", "D", "E", "F", "G", "H"]
        column_indices = [2, 4, 5, 6, 7, 8]  # Skip Column C (index 3)

        # Adding tickmark formulas to columns B to H
        for col, col_index in zip(columns, column_indices):
            # Create the formula for each column
            formula = f"=IF(ROUND(SUM({col}${first_data_row}:{col}{last_data_row})-{col}{total_row},0)=0,\"a\",\"û\")"
            
            # Add the formula to the tickmark row in the respective column
            tickmark_cell = component_sheet.cell(row=tickmark_row, column=col_index, value=formula)
            
            # Apply the formatting (Wingdings font, size 11, centered)
            tickmark_cell.font = Font(name="Wingdings", size=11)
            tickmark_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            logger.info(f"Tickmark formula added to row {tickmark_row}, Column {col} with formula: {formula}")
            
        # Now add the new formula in Column I
        col_i_formula = f"=IF(AND((ROUND(SUM(E{total_row}:G{total_row})-H{total_row},0)=0),ROUND((+B{total_row}+D{total_row})-E{total_row},0)=0),\"b\",\"û\")"
        
        # Add the formula to the tickmark row in Column I
        tickmark_cell_col_i = component_sheet.cell(row=tickmark_row, column=9, value=col_i_formula)
        
        # Apply the formatting (Wingdings font, size 11, centered)
        tickmark_cell_col_i.font = Font(name="Wingdings", size=11)
        tickmark_cell_col_i.alignment = Alignment(horizontal="center", vertical="center")
    
        logger.info(f"Tickmark formula added to row {tickmark_row}, Column I with formula: {col_i_formula}")

        # Now handle the "System of Record vs TIER" tickmark row (inserted or found row)
        system_tickmark_row = system_of_record_row + 1

        # Check if the row under the system of record row is blank, if not, insert a row
        if component_sheet.cell(row=system_tickmark_row, column=2).value is not None:
            component_sheet.insert_rows(system_tickmark_row)
            logger.info(f"Inserted a new row at {system_tickmark_row} for System of Record vs TIER tickmark")
            
            # Increment all the other relevant row references to account for the inserted row
            udo_total_system_row += 1
            udo_after_adjustments_row += 1
            udo_tickmark_row += 1
            difference_adjustments_row += 1
            difference_adjustments_tickmark_row += 1
        
        # Update formulas to use the passed udo_row instead of uco_row
        system_b_formula = (
            f"=IF(AND(B{udo_row+1}=\"i\",B{system_of_record_row}>0),"
            f"IF(ROUND(B{total_row}-B{udo_row}+B{system_of_record_row},0)=0,\"a\",\"û\"),"
            f"IF(ROUND(B{total_row}-B{udo_row}-IF(B{system_of_record_row}<0,-B{system_of_record_row},B{system_of_record_row}),0)=0,\"a\",\"û\"))"
        )
        
        system_d_formula = (
            f"=IF(AND(D{udo_row+1}=\"i\",D{system_of_record_row}>0),"
            f"IF(ROUND(D{total_row}-D{udo_row}+D{system_of_record_row},0)=0,\"a\",\"û\"),"
            f"IF(ROUND(D{total_row}-D{udo_row}-IF(D{system_of_record_row}<0,-D{system_of_record_row},D{system_of_record_row}),0)=0,\"a\",\"û\"))"
        )
        
        # Add the formulas to the System of Record vs TIER tickmark row in Columns B and D
        system_b_cell = component_sheet.cell(row=system_tickmark_row, column=2, value=system_b_formula)
        system_d_cell = component_sheet.cell(row=system_tickmark_row, column=4, value=system_d_formula)
        
        # Apply formatting (Wingdings font, size 11, centered) for both cells
        for cell in [system_b_cell, system_d_cell]:
            cell.font = Font(name="Wingdings", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        logger.info(f"System of Record vs TIER tickmark formulas added to row {system_tickmark_row}, Columns B and D")
        
        # Adding the formula in Column D for UDO After High Level Adjustments formula row if rows are shifted down
        udo_hla_formula = f"=SUM(D{udo_total_system_row}:D{udo_after_adjustments_row-1})"
        udo_hla_formula_cell = component_sheet.cell(row=udo_after_adjustments_row, column=4, value=udo_hla_formula)

        logger.info(f"UDO After High Level Adjustments formula added to row {udo_after_adjustments_row}, Column D with formula: {udo_hla_formula}")
        
        # Adding the formula for "Difference After Adjustments formula" in Column D if rows are shifted down
        difference_adj_formula = f"=+D{udo_after_adjustments_row}-D{udo_row}"
        difference_adj_formula_cell = component_sheet.cell(row=difference_adjustments_row, column=4, value=difference_adj_formula)

        logger.info(f"Difference After Adjustments formula added to row {difference_adjustments_row}, Column D with formula: {difference_adj_formula}")

        # Adding the formula in Column D for UDO After High Level Adjustments tickmark row
        udo_formula = f"=IF(ROUND(SUM(D{udo_total_system_row}:D{udo_after_adjustments_row-1})-D{udo_after_adjustments_row},0)=0,\"a\",\"û\")"
        udo_tickmark_cell = component_sheet.cell(row=udo_tickmark_row, column=4, value=udo_formula)
        
        # Apply formatting (Wingdings font, size 11, centered)
        udo_tickmark_cell.font = Font(name="Wingdings", size=11)
        udo_tickmark_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        logger.info(f"UDO After High Level Adjustments tickmark formula added to row {udo_tickmark_row}, Column D with formula: {udo_formula}")
        
        # New Addition: Add the formula for "Difference After Adjustments Tickmark Row" in Column D
        difference_adjustments_formula = f"=IF(ROUND(+D{udo_row}-D{udo_after_adjustments_row}+D{difference_adjustments_row},0)=0,\"a\",\"û\")"
        difference_adjustments_tickmark_cell = component_sheet.cell(row=difference_adjustments_tickmark_row, column=4, value=difference_adjustments_formula)
        
        # Apply formatting (Wingdings font, size 11, centered)
        difference_adjustments_tickmark_cell.font = Font(name="Wingdings", size=11)
        difference_adjustments_tickmark_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        logger.info(f"Difference After Adjustments tickmark formula added to row {difference_adjustments_tickmark_row}, Column D with formula: {difference_adjustments_formula}")
        
        # Save the workbook with the new_target_file
        component_sheet.parent.save(new_target_file)
        
    except Exception as e:
        logger.error(f"An error occurred while processing the recon table: {e}", exc_info=True)

def add_tickmark(sheet, row, col, value, font_name, font_size, is_match=True):
    """Add a tickmark to a cell with specified formatting."""
    cell = sheet.cell(row=row, column=col, value=value)
    if is_match:
        cell.font = Font(name=font_name, size=font_size)
    else:
        cell.font = Font(name="Calibri", size=font_size, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Apply yellow fill only for the DO TB sheet
    if sheet.title == "DO TB":
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

def convert_to_decimal(value, logger):
    """Convert a value to a Decimal, handling various formats, including accounting format."""
    if value is None:
        logger.warning(f"Encountered None value, converting to Decimal('0')")
        return Decimal('0')
    
    # Handle blank, accounting-format "-", or NaN
    if isinstance(value, str):
        cleaned_value = value.replace(",", "").replace("$", "").strip()
        
        if cleaned_value in ["-", ""]:  # Accounting-style "-" or blank string
            logger.info(f"Detected accounting-format '-' or blank. Treating as Decimal('0').")
            return Decimal('0')
        try:
            if cleaned_value.startswith("(") and cleaned_value.endswith(")"):
                cleaned_value = "-" + cleaned_value[1:-1]
            return Decimal(cleaned_value).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        except InvalidOperation:
            logger.warning(f"Failed to convert string value: {value}")
            return Decimal('0')
    
    # Handle numeric types (int, float)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            logger.warning(f"Encountered NaN value, converting to Decimal('0')")
            return Decimal('0')
        return Decimal(str(value)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    
    logger.warning(f"Unexpected value type: {type(value)}")
    return Decimal('0')

def find_component_sheet(workbook, tab_name, tier_component_name, trading_partner_number, logger):
    """
    Find the component sheet using multiple fallback strategies with comprehensive component mappings
    
    Args:
        workbook: The Excel workbook object
        tab_name: The tab name to search for (can be None)
        tier_component_name: The TIER component name (e.g., 'FEM', 'CBP')
        trading_partner_number: The trading partner number
        logger: Logger instance
    
    Returns:
        worksheet object or None
    """
    try:
        # Complete component mappings based on the Excel sheet
        component_mappings = {
            "CBP": ["CBP", "CBP-7005"],
            "CG": ["USCG", "CG", "USCG-7006"],
            "CIS": ["CIS", "CIS-7001"],
            "CYB": ["CISA", "CYB", "CISA-7009"],
            "FEM": ["FEMA", "FEM", "FEMA-7007"],
            "ICE": ["ICE", "ICE-7019"],
            "MGA": ["MGA", "MGA-7021"],
            "MGT": ["MGT", "MGT-7003"],
            "OIG": ["OIG", "OIG-7002"],
            "SS": ["USSS", "SS", "USSS-7004"],
            "ST": ["ST", "STA-7008"],
            "TSA": ["TSA", "TSA-7011"],
            "WMD": ["CWMD", "WMD", "CWMD-7023"]
        }

        # Trading partner mappings
        trading_partner_mappings = {
            "7005": ["CBP-7005"],
            "7006": ["USCG-7006"],
            "7001": ["CIS-7001"],
            "7009": ["CISA-7009"],
            "7007": ["FEMA-7007"],
            "7019": ["ICE-7019"],
            "7021": ["MGA-7021"],
            "7003": ["MGT-7003"],
            "7002": ["OIG-7002"],
            "7004": ["USSS-7004"],
            "7008": ["STA-7008"],
            "7011": ["TSA-7011"],
            "7023": ["CWMD-7023"]
        }

        # Normalize input values and create search patterns
        search_patterns = []
        
        # Build search patterns, handling None values
        if tab_name is not None:
            search_patterns.append((str(tab_name), "tab_name"))
        
        if tier_component_name is not None:
            # Add all possible component variations
            if tier_component_name in component_mappings:
                for variant in component_mappings[tier_component_name]:
                    search_patterns.append((variant, f"component_mapping_{tier_component_name}"))
            else:
                search_patterns.append((str(tier_component_name), "tier_component_name"))
        
        if trading_partner_number is not None:
            # Add trading partner variations
            str_trading_partner = str(trading_partner_number)
            search_patterns.append((str_trading_partner, "trading_partner_number"))
            if str_trading_partner in trading_partner_mappings:
                for variant in trading_partner_mappings[str_trading_partner]:
                    search_patterns.append((variant, f"trading_partner_mapping_{str_trading_partner}"))

        # List of sheets to skip
        skip_sheets = {
            "Instructions", 
            "Certification", 
            "DO TB", 
            "DO UCO to UDO"
        }

        # Search through all patterns
        for sheet_name in workbook.sheetnames:
            # Skip certain sheet names
            if sheet_name in skip_sheets:
                continue
                
            sheet_name_upper = sheet_name.upper()
            
            for pattern, pattern_type in search_patterns:
                if pattern and pattern.upper() in sheet_name_upper:
                    logger.info(f"Found sheet '{sheet_name}' using {pattern_type} pattern: {pattern}")
                    return workbook[sheet_name]

        # If no match found, log detailed information
        logger.warning(
            f"No matching sheet found for TIER Component: {tier_component_name}\n"
            f"Search details:\n"
            f"- Tab Name: {tab_name}\n"
            f"- Trading Partner: {trading_partner_number}\n"
            f"- Search Patterns Used: {[p[0] for p in search_patterns]}\n"
            f"- Available sheets: {', '.join(workbook.sheetnames)}"
        )
        
        # Additional debugging information
        if tier_component_name:
            logger.debug(f"Component mappings available for {tier_component_name}: "
                        f"{component_mappings.get(tier_component_name, 'None')}")
        if trading_partner_number:
            logger.debug(f"Trading partner mappings available for {trading_partner_number}: "
                        f"{trading_partner_mappings.get(str(trading_partner_number), 'None')}")

        return None

    except Exception as e:
        logger.error(
            f"Error in find_component_sheet:\n"
            f"- Error message: {str(e)}\n"
            f"- TIER Component: {tier_component_name}\n"
            f"- Tab Name: {tab_name}\n"
            f"- Trading Partner: {trading_partner_number}\n"
            f"- Available sheets: {', '.join(workbook.sheetnames)}",
            exc_info=True
        )
        return None

def main(certification_range, uco_to_udo_range, target_wb, data_wb, logger, progress_callback, new_target_file, cancel_event=None):
    """
    Main function to run the comparison process for UCO and UDO values.

    Args:
        certification_range (list of tuples): The range of cells from the Certification sheet.
        uco_to_udo_range (list of tuples): The range of cells from the DO UCO to UDO sheet.
        target_wb (Workbook): The workbook loaded with data_only=False (preserving formulas).
        data_wb (Workbook): The workbook loaded with data_only=True (accessing calculated values).
        logger (Logger): Logger instance for logging.
        progress_callback (function): Callback function to update progress.
        new_target_file (str): Path to save the updated workbook.
        cancel_event (threading.Event, optional): Event to check for cancellation.
    """
    # Check for cancellation
    if cancel_event and cancel_event.is_set():
        logger.info("Operation cancelled before comparison.")
        return

    # Execute the comparison
    compare_ranges(certification_range, uco_to_udo_range, target_wb, data_wb, logger, progress_callback, new_target_file, cancel_event)


if __name__ == "__main__":
    # This script is intended to be imported and run via other scripts.
    logging.basicConfig(level=logging.DEBUG)
    logger = logging.getLogger("CompareRangesLogger")
    # main(certification_range, uco_to_udo_range, workbook, uco_to_udo_sheet, certification_row_data, logger, progress_callback)  # This should be called with actual ranges, workbook, and logger
